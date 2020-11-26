import getpass
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import openpyxl


def scrape(url: str, item_list: list):
    print(item_list)
    output = []
    if url == "":
        for item in item_list:
            if not bool(item):
                continue
            print(item)
            output.append(item)
        return output

    driver_path = "C:\\Users\\" + getpass.getuser() + "\\webdriver\\chromedriver.exe"
    driver = webdriver.Chrome(driver_path)
    driver.get(url)
    input_text = driver.find_element_by_id("ENTRY3")
    for item in item_list:
        if not bool(item):
            continue
        print(item)
        input_text.send_keys(item)
        input_text.send_keys(Keys.RETURN)
        time.sleep(0.5)
        on_hand_qty = driver.find_element_by_id("OUTPUT55")
        commit_qty = driver.find_element_by_id("OUTPUT57")
        avail_qty = driver.find_element_by_id("OUTPUT59")
        backorder_qty = driver.find_element_by_id("OUTPUT61")
        open_order_qty = driver.find_element_by_id("OUTPUT62")
        item_status = driver.find_element_by_id("OUTPUT3")
        item_info = get_info_by_item_status(
            int(avail_qty), int(backorder_qty), int(open_order_qty)
        )
        output.append(
            (
                str(item),
                avail_qty,
                on_hand_qty,
                commit_qty,
                backorder_qty,
                open_order_qty,
                item_status,
                item_info[0],
                item_info[1],
            )
        )
    return output


def get_info_by_item_status(avail_qty: int, bo_qty: int, po_qty: int):
    notes = ""
    contact = ""
    if avail_qty + po_qty == 0:
        notes += "Order In"
    elif po_qty == 1:
        notes += "Order More?"

    if avail_qty == 0 and bo_qty > 0:
        notes += "Transfer In?"

    if notes != "":
        contact += "Kari"

    if avail_qty > 0:
        contact += "Mark"
        notes = "check into"

    return notes, contact


def list_to_excel(item_list):
    desktop_path = "C:\\Users\\" + getpass.getuser() + "\\Desktop\\"
    print(desktop_path)
    wb = openpyxl.Workbook()
    dot_sheet = wb.active
    header = [
        "Item Number",
        "# Avail",
        "# Commit",
        "# OH",
        "# BO",
        "# PO",
        "Status",
        "Notes",
        "Contact",
    ]
    dim_dict = {
        "A": 18,
        "B": 7,
        "C": 9,
        "D": 5,
        "E": 5,
        "F": 5,
        "G": 10,
        "H": 21,
        "I": 7,
    }

    row = 1
    col = 1
    for enum in header:
        dot_sheet.cell(row, col).value = enum
        col += 1

    row = 2
    col = 1
    for item in item_list:
        for i in item:
            dot_sheet.cell(row, col).value = i
            col += 1
        col = 1
        row += 1

    for col in dim_dict:
        dot_sheet.column_dimensions[col].width = dim_dict[col]

    current_time = time.strftime("%m-%d-%y")
    wb.save(desktop_path + "Dot-Sheet-" + current_time + ".xlsx")
